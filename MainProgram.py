import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar
from datetime import datetime


class ExcelMacroConverter:
    def __init__(self, workbook_path):
        """Initialize with Excel workbook path"""
        self.workbook_path = workbook_path
        self.workbook = None
        self.load_workbook()

    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.workbook_path, data_only=False)
        except FileNotFoundError:
            print(f"File not found: {self.workbook_path}")
        except Exception as e:
            print(f"Error loading workbook: {e}")

    def save_workbook(self):
        """Save the workbook"""
        if self.workbook:
            self.workbook.save(self.workbook_path)

    def macro10(self, sheet_name=None):
        """
        Macro10: Refresh all data and delete first row
        VBA equivalent:
        - ActiveWorkbook.RefreshAll
        - Rows("1:1").Select -> Selection.Delete Shift:=xlUp
        """
        try:
            if sheet_name:
                ws = self.workbook[sheet_name]
            else:
                ws = self.workbook.active

            # Note: RefreshAll is not directly equivalent in openpyxl
            # For data refresh, you might need to implement specific logic
            print("RefreshAll functionality needs to be implemented based on data sources")

            # Delete first row (equivalent to Rows("1:1").Delete Shift:=xlUp)
            ws.delete_rows(1, 1)

            print("Macro10 completed: First row deleted")

        except Exception as e:
            print(f"Error in macro10: {e}")

    def macro7(self, sheet_name="MainReport"):
        """
        Macro7: Sort MainReport sheet by column B in descending order
        VBA equivalent: Sort range A1:AO171000 by column B descending
        """
        try:
            ws = self.workbook[sheet_name]

            # Get data range - find last row and column with data
            max_row = min(171000, ws.max_row)
            max_col = min(41, ws.max_column)  # AO is column 41

            # Convert worksheet data to pandas DataFrame for sorting
            data = []
            headers = []

            # Get headers from row 1
            for col in range(1, max_col + 1):
                headers.append(ws.cell(row=1, column=col).value)

            # Get data from rows 2 onwards
            for row in range(2, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                data.append(row_data)

            # Create DataFrame and sort by column B (index 1) in descending order
            df = pd.DataFrame(data, columns=headers)
            df = df.sort_values(by=df.columns[1], ascending=False, na_position='last')

            # Clear the worksheet data area
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).value = None

            # Write sorted data back to worksheet
            for idx, row_data in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=idx, column=col_idx).value = value

            print("Macro7 completed: Data sorted by column B (descending)")

        except Exception as e:
            print(f"Error in macro7: {e}")

    def macro8(self, sheet_name=None):
        """
        Macro8: Delete rows from row 44 to end of data
        VBA equivalent: Rows("44:44").Select -> Range(Selection, Selection.End(xlDown)).Select -> Selection.Delete
        """
        try:
            if sheet_name:
                ws = self.workbook[sheet_name]
            else:
                ws = self.workbook.active

            # Find the last row with data starting from row 44
            start_row = 44
            last_row = ws.max_row

            if last_row >= start_row:
                # Delete rows from 44 to last row
                ws.delete_rows(start_row, last_row - start_row + 1)
                print(f"Macro8 completed: Deleted rows {start_row} to {last_row}")
            else:
                print("Macro8: No rows to delete")

        except Exception as e:
            print(f"Error in macro8: {e}")

    def macro9(self, sheet_name="MainReport"):
        """
        Macro9: Multi-level sort by column E (ascending) then column B (descending)
        VBA equivalent: Sort by column E ascending, then by column B descending
        """
        try:
            ws = self.workbook[sheet_name]

            # Get data range
            max_row = min(225000, ws.max_row)
            max_col = min(41, ws.max_column)  # AO is column 41

            # Convert to DataFrame
            data = []
            headers = []

            # Get headers
            for col in range(1, max_col + 1):
                headers.append(ws.cell(row=1, column=col).value)

            # Get data
            for row in range(2, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                data.append(row_data)

            # Create DataFrame and sort by column E (index 4) ascending, then column B (index 1) descending
            df = pd.DataFrame(data, columns=headers)
            df = df.sort_values(by=[df.columns[4], df.columns[1]],
                                ascending=[True, False], na_position='last')

            # Clear and write back sorted data
            for row in range(2, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).value = None

            for idx, row_data in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=idx, column=col_idx).value = value

            print("Macro9 completed: Data sorted by column E (asc) then column B (desc)")

        except Exception as e:
            print(f"Error in macro9: {e}")

    def macro11(self, sheet_name=None, end_row=89):
        """
        Macro11: Add formulas in columns C and D, then convert to values
        VBA equivalent:
        - C2: =TEXT(29*MONTH(B2),"mmm")
        - D2: =DAY(B2)
        - AutoFill to row 89, convert to values
        """
        try:
            if sheet_name:
                ws = self.workbook[sheet_name]
            else:
                ws = self.workbook.active

            # Process each row from 2 to end_row
            for row in range(2, end_row + 1):
                # Get the date value from column B
                date_value = ws.cell(row=row, column=2).value

                if date_value is not None:
                    try:
                        # Convert to datetime if it's not already
                        if isinstance(date_value, str):
                            date_value = pd.to_datetime(date_value)
                        elif isinstance(date_value, (int, float)):
                            # Excel date serial number to datetime
                            date_value = datetime(1900, 1, 1) + pd.Timedelta(days=date_value - 2)

                        # Column C: Month abbreviation (TEXT(29*MONTH(date),"mmm"))
                        month_num = date_value.month
                        month_abbr = calendar.month_abbr[month_num]
                        ws.cell(row=row, column=3).value = month_abbr

                        # Column D: Day of month
                        day_num = date_value.day
                        ws.cell(row=row, column=4).value = day_num

                    except Exception as date_error:
                        print(f"Error processing date in row {row}: {date_error}")
                        continue

            # Set number format for column D to integer (equivalent to "0")
            for row in range(2, end_row + 1):
                cell = ws.cell(row=row, column=4)
                if cell.value is not None:
                    cell.number_format = '0'

            print(f"Macro11 completed: Formulas applied and converted to values (rows 2-{end_row})")

        except Exception as e:
            print(f"Error in macro11: {e}")

    def macro1(self, sheet_name=None):
        """
        Macro1: Remove duplicates from columns B:AO based on columns B and E (indices 1 and 4)
        VBA equivalent: RemoveDuplicates Columns:=Array(1, 4)
        """
        try:
            if sheet_name:
                ws = self.workbook[sheet_name]
            else:
                ws = self.workbook.active

            # Get data range B1:AO127
            start_col = 2  # Column B
            end_col = 41  # Column AO
            end_row = min(127, ws.max_row)

            # Convert to DataFrame
            data = []
            headers = []

            # Get headers from row 1
            for col in range(start_col, end_col + 1):
                headers.append(ws.cell(row=1, column=col).value)

            # Get data from rows 2 onwards
            for row in range(2, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                data.append(row_data)

            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Remove duplicates based on first column (B) and fourth column (E)
            # In the array [1,4], these refer to the 1st and 4th columns in the B:AO range
            subset_cols = [df.columns[0], df.columns[3]]  # Columns B and E
            df_unique = df.drop_duplicates(subset=subset_cols, keep='first')

            # Clear the data area
            for row in range(2, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws.cell(row=row, column=col).value = None

            # Write unique data back
            for idx, row_data in enumerate(df_unique.values, start=2):
                for col_idx, value in enumerate(row_data):
                    ws.cell(row=idx, column=col_idx + start_col).value = value

            original_count = len(df)
            unique_count = len(df_unique)
            removed_count = original_count - unique_count

            print(f"Macro1 completed: Removed {removed_count} duplicate rows (kept {unique_count} unique rows)")

        except Exception as e:
            print(f"Error in macro1: {e}")


# Example usage
def main():
    """Example of how to use the ExcelMacroConverter class"""

    # Initialize with your Excel file path
    converter = ExcelMacroConverter("4G_KPI Cell FDD Data_24h_scheduled.xlsx")

    # Execute the macros in sequence (adjust as needed for your workflow)
    try:
        converter.macro10()  # Refresh and delete first row
        converter.macro7("MainReport")  # Sort by column B descending
        converter.macro8()  # Delete rows from 44 onwards
        converter.macro9("MainReport")  # Multi-level sort
        converter.macro11()  # Add date formulas
        converter.macro1()  # Remove duplicates

        # Save the workbook after all operations
        converter.save_workbook()
        print("All macros completed successfully!")

    except Exception as e:
        print(f"Error in main execution: {e}")


if __name__ == "__main__":
    # Uncomment the line below to run the example
    main()

    # Or run individual macros as needed:
    converter = ExcelMacroConverter("4G_KPI Cell FDD Data_24h_scheduled.xlsx")
    converter.macro7("MainReport")
    converter.save_workbook()

    print("VBA to Python conversion completed. Use ExcelMacroConverter class to execute macros.")